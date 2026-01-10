"""
Unit tests for product import/export functionality.
"""
import io
import csv
import pytest
import openpyxl

from apps.products.models import Product, Listing
from apps.products.import_export import ProductImporter, ProductExporter, ImportResult
from apps.retailers.models import Retailer


@pytest.fixture
def retailers(db):
    """Create test retailers."""
    ozon = Retailer.objects.create(
        name='Ozon',
        slug='ozon',
        connector_class='apps.scraping.connectors.ozon.OzonConnector',
        base_url='https://ozon.ru',
    )
    vkusvill = Retailer.objects.create(
        name='ВкусВилл',
        slug='vkusvill',
        connector_class='apps.scraping.connectors.vkusvill.VkusvillConnector',
        base_url='https://vkusvill.ru',
    )
    return {'ozon': ozon, 'vkusvill': vkusvill}


class TestProductImporter:
    """Tests for ProductImporter."""

    def test_normalize_column_russian(self):
        importer = ProductImporter()
        assert importer._normalize_column('Название') == 'name'
        assert importer._normalize_column('бренд') == 'brand'
        assert importer._normalize_column('ВЕС (Г)') == 'weight_grams'

    def test_normalize_column_english(self):
        importer = ProductImporter()
        assert importer._normalize_column('name') == 'name'
        assert importer._normalize_column('brand') == 'brand'
        assert importer._normalize_column('url_ozon') == 'url_ozon'

    def test_normalize_column_unknown(self):
        importer = ProductImporter()
        assert importer._normalize_column('unknown_column') is None
        assert importer._normalize_column('foo') is None

    def test_parse_bool_russian(self):
        importer = ProductImporter()
        assert importer._parse_bool('Да') is True
        assert importer._parse_bool('Нет') is False
        assert importer._parse_bool('да') is True
        assert importer._parse_bool('нет') is False

    def test_parse_bool_english(self):
        importer = ProductImporter()
        assert importer._parse_bool('yes') is True
        assert importer._parse_bool('no') is False
        assert importer._parse_bool('true') is True
        assert importer._parse_bool('false') is False

    def test_parse_bool_numeric(self):
        importer = ProductImporter()
        assert importer._parse_bool('1') is True
        assert importer._parse_bool('0') is False
        assert importer._parse_bool('+') is True
        assert importer._parse_bool('-') is False

    def test_parse_bool_empty(self):
        importer = ProductImporter()
        assert importer._parse_bool('') is None
        assert importer._parse_bool(None) is None

    def test_parse_int(self):
        importer = ProductImporter()
        assert importer._parse_int('500') == 500
        assert importer._parse_int('1 000') == 1000
        assert importer._parse_int('250г') == 250

    def test_parse_int_empty(self):
        importer = ProductImporter()
        assert importer._parse_int('') is None
        assert importer._parse_int(None) is None

    def test_parse_packaging(self):
        importer = ProductImporter()
        assert importer._parse_packaging('дой-пак') == 'doypack'
        assert importer._parse_packaging('Дойпак') == 'doypack'
        assert importer._parse_packaging('box') == 'box'
        assert importer._parse_packaging('коробка') == 'box'
        assert importer._parse_packaging('unknown') == ''

    @pytest.mark.django_db
    def test_import_xlsx_creates_products(self, retailers):
        """Test importing products from Excel."""
        # Create a test Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Название', 'Бренд', 'Наш товар', 'Ozon'])
        ws.append(['Test Kuraga', 'TestBrand', 'Да', 'https://ozon.ru/product/test-123/'])
        ws.append(['Test Chernosliv', 'TestBrand', 'Нет', ''])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        importer = ProductImporter()
        result = importer.import_xlsx(buffer)

        assert result.total_rows == 2
        assert result.products_created == 2
        assert result.listings_created == 1
        assert result.success is True

        # Verify products
        kuraga = Product.objects.get(name='Test Kuraga')
        assert kuraga.brand == 'TestBrand'
        assert kuraga.is_own is True
        assert kuraga.listings.count() == 1
        assert kuraga.listings.first().retailer.slug == 'ozon'

        chernosliv = Product.objects.get(name='Test Chernosliv')
        assert chernosliv.is_own is False
        assert chernosliv.listings.count() == 0

    @pytest.mark.django_db
    def test_import_xlsx_updates_existing(self, retailers):
        """Test that import updates existing products."""
        # Create existing product
        Product.objects.create(name='Existing Product', brand='OldBrand', is_own=True)

        # Create Excel with updated data
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Название', 'Бренд', 'Тип продукта'])
        ws.append(['Existing Product', 'OldBrand', 'Сухофрукты'])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        importer = ProductImporter()
        result = importer.import_xlsx(buffer)

        assert result.products_created == 0
        assert result.products_updated == 1

        product = Product.objects.get(name='Existing Product')
        assert product.product_type == 'Сухофрукты'

    @pytest.mark.django_db
    def test_import_xlsx_missing_required_columns(self):
        """Test import fails without required columns."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Тип', 'Цена'])
        ws.append(['Сухофрукты', '500'])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        importer = ProductImporter()
        result = importer.import_xlsx(buffer)

        assert len(result.errors) > 0
        assert 'обязательные колонки' in result.errors[0].lower()

    @pytest.mark.django_db
    def test_import_csv_creates_products(self, retailers):
        """Test importing products from CSV."""
        csv_content = 'Название,Бренд,Наш товар,Ozon\n'
        csv_content += 'CSV Kuraga,CSVBrand,Да,https://ozon.ru/product/csv-123/\n'

        buffer = io.BytesIO(csv_content.encode('utf-8'))

        importer = ProductImporter()
        result = importer.import_csv(buffer)

        assert result.products_created == 1
        assert result.listings_created == 1

        product = Product.objects.get(name='CSV Kuraga')
        assert product.brand == 'CSVBrand'

    @pytest.mark.django_db
    def test_import_csv_windows_encoding(self, retailers):
        """Test importing CSV with Windows-1251 encoding."""
        csv_content = 'Название,Бренд\nТест Товар,Тест Бренд\n'

        buffer = io.BytesIO(csv_content.encode('cp1251'))

        importer = ProductImporter()
        result = importer.import_csv(buffer)

        assert result.products_created == 1
        product = Product.objects.get(name='Тест Товар')
        assert product.brand == 'Тест Бренд'


@pytest.mark.django_db
class TestProductExporter:
    """Tests for ProductExporter."""

    @pytest.fixture
    def products_with_listings(self, retailers):
        """Create products with listings for export tests."""
        p1 = Product.objects.create(
            name='Export Product 1',
            brand='ExportBrand',
            is_own=True,
            product_type='Сухофрукты',
            weight_grams=500,
        )
        Listing.objects.create(
            product=p1,
            retailer=retailers['ozon'],
            external_url='https://ozon.ru/product/export-1/',
        )

        p2 = Product.objects.create(
            name='Export Product 2',
            brand='ExportBrand',
            is_own=False,
        )
        return [p1, p2]

    def test_export_xlsx(self, products_with_listings):
        """Test exporting products to Excel."""
        exporter = ProductExporter()
        content = exporter.export_xlsx()

        # Load and verify
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert 'Название' in headers
        assert 'Бренд' in headers
        assert 'URL Ozon' in headers

        # Check data rows (skip header)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 2

        # Find the product row
        names = [row[0] for row in rows]
        assert 'Export Product 1' in names
        assert 'Export Product 2' in names

    def test_export_csv(self, products_with_listings):
        """Test exporting products to CSV."""
        exporter = ProductExporter()
        content = exporter.export_csv()

        # Parse CSV
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        assert len(rows) == 3  # Header + 2 products

        # Check headers
        assert 'Название' in rows[0]
        assert 'Бренд' in rows[0]

        # Check data
        names = [row[0] for row in rows[1:]]
        assert 'Export Product 1' in names

    def test_export_filtered_queryset(self, products_with_listings):
        """Test exporting filtered queryset."""
        exporter = ProductExporter()
        content = exporter.export_xlsx(Product.objects.filter(is_own=True))

        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1
        assert rows[0][0] == 'Export Product 1'


class TestImportResult:
    """Tests for ImportResult dataclass."""

    def test_success_when_no_errors(self):
        result = ImportResult(products_created=5)
        assert result.success is True

    def test_not_success_when_errors(self):
        result = ImportResult()
        result.add_error(1, 'Test error')
        assert result.success is False

    def test_add_error_formats_correctly(self):
        result = ImportResult()
        result.add_error(5, 'Missing required field')
        assert result.errors[0] == 'Строка 5: Missing required field'

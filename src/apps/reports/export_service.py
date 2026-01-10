"""
Report export service - generates XLSX reports with multiple sheets.
"""
import io
import logging
from datetime import date, datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Q

logger = logging.getLogger(__name__)


class ReportExporter:
    """
    Generates XLSX reports with price matrix, reviews matrix, and insights.
    """

    # Style definitions
    HEADER_FONT = Font(bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

    SUBHEADER_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    OWN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    COMPETITOR_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    NEGATIVE_FILL = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    POSITIVE_FILL = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')

    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    def __init__(
        self,
        period_from: date = None,
        period_to: date = None,
        is_own: Optional[bool] = None,
        retailer_id: Optional[str] = None,
    ):
        """
        Initialize exporter with filters.

        Args:
            period_from: Start date (default: 6 months ago)
            period_to: End date (default: today)
            is_own: Filter by own products (None = all)
            retailer_id: Filter by retailer
        """
        from dateutil.relativedelta import relativedelta

        self.period_to = period_to or date.today()
        self.period_from = period_from or (self.period_to - relativedelta(months=6))
        self.is_own = is_own
        self.retailer_id = retailer_id

    def generate_full_report(self) -> bytes:
        """
        Generate full report with all 3 sheets.

        Returns:
            XLSX file as bytes
        """
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        self._create_price_matrix_sheet(wb)
        self._create_reviews_matrix_sheet(wb)
        self._create_insights_sheet(wb)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def generate_price_matrix(self) -> bytes:
        """Generate price matrix report only."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        self._create_price_matrix_sheet(wb)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def generate_reviews_matrix(self) -> bytes:
        """Generate reviews matrix report only."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        self._create_reviews_matrix_sheet(wb)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def generate_insights_report(self) -> bytes:
        """Generate insights report only."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        self._create_insights_sheet(wb)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def _get_months_range(self) -> list[date]:
        """Get list of months in the period."""
        from dateutil.relativedelta import relativedelta

        months = []
        current = self.period_from.replace(day=1)
        end = self.period_to.replace(day=1)

        while current <= end:
            months.append(current)
            current += relativedelta(months=1)

        return months

    def _get_products_queryset(self):
        """Get filtered products queryset."""
        from apps.products.models import Product

        queryset = Product.objects.prefetch_related(
            'listings__retailer',
            'listings__price_snapshots',
        ).order_by('is_own', 'brand', 'name')

        if self.is_own is not None:
            queryset = queryset.filter(is_own=self.is_own)

        if self.retailer_id:
            queryset = queryset.filter(listings__retailer_id=self.retailer_id).distinct()

        return queryset

    def _create_price_matrix_sheet(self, wb: openpyxl.Workbook):
        """
        Create price matrix sheet.
        Rows: Products
        Columns: Months x Retailers (price_final)
        """
        from apps.retailers.models import Retailer
        from apps.scraping.models import SnapshotPrice

        ws = wb.create_sheet('Цены')

        # Get data
        products = list(self._get_products_queryset())
        months = self._get_months_range()
        retailers = list(Retailer.objects.filter(is_active=True).order_by('name'))

        if not products:
            ws['A1'] = 'Нет данных'
            return

        # Headers
        # Row 1: Month names spanning retailer columns
        # Row 2: Retailer names
        ws.cell(row=1, column=1, value='Товар')
        ws.cell(row=2, column=1, value='')
        ws.cell(row=1, column=2, value='Бренд')
        ws.cell(row=2, column=2, value='')
        ws.cell(row=1, column=3, value='Тип')
        ws.cell(row=2, column=3, value='')

        col = 4
        for month in months:
            month_str = month.strftime('%b %Y')
            ws.cell(row=1, column=col, value=month_str)
            ws.merge_cells(
                start_row=1, start_column=col,
                end_row=1, end_column=col + len(retailers) - 1
            )

            for retailer in retailers:
                cell = ws.cell(row=2, column=col, value=retailer.name)
                cell.fill = self.SUBHEADER_FILL
                cell.alignment = self.HEADER_ALIGNMENT
                col += 1

        # Apply header styles
        for c in range(1, col):
            cell = ws.cell(row=1, column=c)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT

        # Data rows
        row_num = 3
        for product in products:
            ws.cell(row=row_num, column=1, value=product.name)
            ws.cell(row=row_num, column=2, value=product.brand)
            ws.cell(row=row_num, column=3, value='Наш' if product.is_own else 'Конкурент')

            # Apply row color
            row_fill = self.OWN_FILL if product.is_own else self.COMPETITOR_FILL
            for c in range(1, 4):
                ws.cell(row=row_num, column=c).fill = row_fill

            col = 4
            for month in months:
                for retailer in retailers:
                    # Find listing for this product+retailer
                    listing = next(
                        (l for l in product.listings.all() if l.retailer_id == retailer.id),
                        None
                    )

                    if listing:
                        # Get price snapshot for this month
                        snapshot = SnapshotPrice.objects.filter(
                            listing=listing,
                            period_month=month,
                        ).order_by('-scraped_at').first()

                        if snapshot and snapshot.price_final:
                            cell = ws.cell(row=row_num, column=col, value=float(snapshot.price_final))
                            cell.number_format = '#,##0.00 ₽'
                        else:
                            ws.cell(row=row_num, column=col, value='—')
                    else:
                        ws.cell(row=row_num, column=col, value='')

                    col += 1

            row_num += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        for c in range(4, col):
            ws.column_dimensions[get_column_letter(c)].width = 12

    def _create_reviews_matrix_sheet(self, wb: openpyxl.Workbook):
        """
        Create reviews matrix sheet.
        Rows: Products
        Columns: Months (rating_avg, reviews_count, negative_count)
        """
        from apps.scraping.models import SnapshotReview

        ws = wb.create_sheet('Отзывы')

        products = list(self._get_products_queryset())
        months = self._get_months_range()

        if not products:
            ws['A1'] = 'Нет данных'
            return

        # Headers
        ws.cell(row=1, column=1, value='Товар')
        ws.cell(row=2, column=1, value='')
        ws.cell(row=1, column=2, value='Бренд')
        ws.cell(row=2, column=2, value='')
        ws.cell(row=1, column=3, value='Тип')
        ws.cell(row=2, column=3, value='')

        col = 4
        for month in months:
            month_str = month.strftime('%b %Y')
            ws.cell(row=1, column=col, value=month_str)
            ws.merge_cells(
                start_row=1, start_column=col,
                end_row=1, end_column=col + 2
            )

            # Sub-headers
            ws.cell(row=2, column=col, value='Рейтинг').fill = self.SUBHEADER_FILL
            ws.cell(row=2, column=col + 1, value='Всего').fill = self.SUBHEADER_FILL
            ws.cell(row=2, column=col + 2, value='Негат.').fill = self.SUBHEADER_FILL
            col += 3

        # Apply header styles
        for c in range(1, col):
            cell = ws.cell(row=1, column=c)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT

        # Data rows
        row_num = 3
        for product in products:
            ws.cell(row=row_num, column=1, value=product.name)
            ws.cell(row=row_num, column=2, value=product.brand)
            ws.cell(row=row_num, column=3, value='Наш' if product.is_own else 'Конкурент')

            listing_ids = list(product.listings.values_list('id', flat=True))

            col = 4
            for month in months:
                # Aggregate review snapshots for all listings
                snapshots = SnapshotReview.objects.filter(
                    listing_id__in=listing_ids,
                    period_month=month,
                )

                if snapshots.exists():
                    total_reviews = sum(
                        s.reviews_1_count + s.reviews_2_count + s.reviews_3_count +
                        s.reviews_4_count + s.reviews_5_count
                        for s in snapshots
                    )
                    negative = sum(s.reviews_1_3_count for s in snapshots)

                    # Calculate average rating
                    weighted_sum = sum(
                        s.reviews_1_count * 1 + s.reviews_2_count * 2 +
                        s.reviews_3_count * 3 + s.reviews_4_count * 4 +
                        s.reviews_5_count * 5
                        for s in snapshots
                    )
                    avg_rating = weighted_sum / total_reviews if total_reviews > 0 else None

                    if avg_rating:
                        ws.cell(row=row_num, column=col, value=round(avg_rating, 1))
                    else:
                        ws.cell(row=row_num, column=col, value='—')

                    ws.cell(row=row_num, column=col + 1, value=total_reviews)

                    neg_cell = ws.cell(row=row_num, column=col + 2, value=negative)
                    if negative > 0:
                        neg_cell.fill = self.NEGATIVE_FILL
                else:
                    ws.cell(row=row_num, column=col, value='—')
                    ws.cell(row=row_num, column=col + 1, value='—')
                    ws.cell(row=row_num, column=col + 2, value='—')

                col += 3

            row_num += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        for c in range(4, col):
            ws.column_dimensions[get_column_letter(c)].width = 10

    def _create_insights_sheet(self, wb: openpyxl.Workbook):
        """
        Create insights sheet with LLM-generated analysis.
        """
        from apps.analytics.models import ReviewAnalysis

        ws = wb.create_sheet('Выводы')

        products = list(self._get_products_queryset())

        if not products:
            ws['A1'] = 'Нет данных'
            return

        # Headers
        headers = [
            'Товар', 'Бренд', 'Тип', 'Ретейлер',
            'Что убрать', 'Изменить упаковку', 'Изменить вкус',
            'Позитивные темы', 'Негативные темы',
            'Инсайты конкурента'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT

        # Data rows
        row_num = 2
        for product in products:
            for listing in product.listings.all():
                # Get latest analysis
                analysis = ReviewAnalysis.objects.filter(
                    listing=listing,
                ).order_by('-period_month').first()

                ws.cell(row=row_num, column=1, value=product.name)
                ws.cell(row=row_num, column=2, value=product.brand)
                ws.cell(row=row_num, column=3, value='Наш' if product.is_own else 'Конкурент')
                ws.cell(row=row_num, column=4, value=listing.retailer.name)

                if analysis:
                    ws.cell(row=row_num, column=5, value=analysis.remove_suggestions)
                    ws.cell(row=row_num, column=6, value=analysis.add_packaging_suggestions)
                    ws.cell(row=row_num, column=7, value=analysis.add_taste_suggestions)
                    ws.cell(row=row_num, column=8, value=', '.join(analysis.key_positive_themes or []))
                    ws.cell(row=row_num, column=9, value=', '.join(analysis.key_negative_themes or []))
                    ws.cell(row=row_num, column=10, value=analysis.competitor_insights)
                else:
                    for col in range(5, 11):
                        ws.cell(row=row_num, column=col, value='Не проанализировано')

                row_num += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        for col in 'EFGHIJ':
            ws.column_dimensions[col].width = 35

        # Enable text wrapping for insight columns
        for row in range(2, row_num):
            for col in range(5, 11):
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

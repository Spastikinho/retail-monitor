"""
Import/Export functionality for products.
"""
import csv
import io
import logging
from dataclasses import dataclass
from typing import Optional, BinaryIO

import openpyxl
from django.db import transaction

from apps.retailers.models import Retailer
from .models import Product, Listing

logger = logging.getLogger(__name__)


@dataclass
class ImportResult:
    """Result of import operation."""
    total_rows: int = 0
    products_created: int = 0
    products_updated: int = 0
    listings_created: int = 0
    listings_updated: int = 0
    errors: list = None

    def __post_init__(self):
        if self.errors is None:
            self.errors = []

    @property
    def success(self):
        return len(self.errors) == 0

    def add_error(self, row: int, message: str):
        self.errors.append(f'Строка {row}: {message}')


class ProductImporter:
    """
    Import products from Excel/CSV files.

    Expected columns:
    - name (required): Product name
    - brand (required): Brand name
    - is_own: "Да"/"Нет" or "1"/"0" or "true"/"false"
    - product_type: Product category
    - packaging_type: doypack, box, bag, tray, jar, other
    - weight_grams: Weight in grams (integer)
    - caliber: Size/caliber
    - has_pit: "Да"/"Нет" - has pit/seed
    - variety: Variety/sort
    - notes: Internal notes
    - url_ozon: Ozon product URL
    - url_vkusvill: VkusVill product URL
    - url_perekrestok: Perekrestok product URL
    - url_lavka: Yandex Lavka product URL
    """

    COLUMN_MAPPING = {
        'название': 'name',
        'name': 'name',
        'наименование': 'name',
        'товар': 'name',

        'бренд': 'brand',
        'brand': 'brand',
        'марка': 'brand',

        'наш': 'is_own',
        'is_own': 'is_own',
        'свой': 'is_own',
        'наш товар': 'is_own',

        'тип': 'product_type',
        'product_type': 'product_type',
        'тип продукта': 'product_type',
        'категория': 'product_type',

        'упаковка': 'packaging_type',
        'packaging_type': 'packaging_type',
        'тип упаковки': 'packaging_type',

        'вес': 'weight_grams',
        'weight_grams': 'weight_grams',
        'вес (г)': 'weight_grams',
        'вес г': 'weight_grams',

        'калибр': 'caliber',
        'caliber': 'caliber',
        'размер': 'caliber',

        'косточка': 'has_pit',
        'has_pit': 'has_pit',
        'с косточкой': 'has_pit',

        'сорт': 'variety',
        'variety': 'variety',

        'заметки': 'notes',
        'notes': 'notes',
        'примечания': 'notes',

        'ozon': 'url_ozon',
        'url_ozon': 'url_ozon',
        'ссылка ozon': 'url_ozon',

        'вкусвилл': 'url_vkusvill',
        'url_vkusvill': 'url_vkusvill',
        'ссылка вкусвилл': 'url_vkusvill',

        'перекресток': 'url_perekrestok',
        'перекрёсток': 'url_perekrestok',
        'url_perekrestok': 'url_perekrestok',
        'ссылка перекресток': 'url_perekrestok',

        'лавка': 'url_lavka',
        'яндекс лавка': 'url_lavka',
        'url_lavka': 'url_lavka',
        'ссылка лавка': 'url_lavka',
    }

    PACKAGING_MAPPING = {
        'дой-пак': 'doypack',
        'дойпак': 'doypack',
        'doypack': 'doypack',
        'коробка': 'box',
        'box': 'box',
        'пакет': 'bag',
        'bag': 'bag',
        'лоток': 'tray',
        'tray': 'tray',
        'банка': 'jar',
        'jar': 'jar',
        'другое': 'other',
        'other': 'other',
    }

    def __init__(self):
        self._retailers = {}

    def _get_retailer(self, slug: str) -> Optional[Retailer]:
        """Get retailer by slug, cached."""
        if slug not in self._retailers:
            try:
                self._retailers[slug] = Retailer.objects.get(slug=slug)
            except Retailer.DoesNotExist:
                self._retailers[slug] = None
        return self._retailers[slug]

    def _normalize_column(self, col: str) -> Optional[str]:
        """Normalize column name to internal field."""
        col_lower = col.strip().lower()
        return self.COLUMN_MAPPING.get(col_lower)

    def _parse_bool(self, value: str) -> Optional[bool]:
        """Parse boolean from string."""
        if not value:
            return None
        value_lower = str(value).strip().lower()
        if value_lower in ('да', 'yes', 'true', '1', '+'):
            return True
        if value_lower in ('нет', 'no', 'false', '0', '-'):
            return False
        return None

    def _parse_int(self, value: str) -> Optional[int]:
        """Parse integer from string."""
        if not value:
            return None
        try:
            # Remove non-digits
            cleaned = ''.join(c for c in str(value) if c.isdigit())
            return int(cleaned) if cleaned else None
        except (ValueError, TypeError):
            return None

    def _parse_packaging(self, value: str) -> str:
        """Parse packaging type."""
        if not value:
            return ''
        value_lower = str(value).strip().lower()
        return self.PACKAGING_MAPPING.get(value_lower, '')

    def import_xlsx(self, file: BinaryIO) -> ImportResult:
        """Import products from Excel file."""
        result = ImportResult()

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            sheet = wb.active

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                result.add_error(0, 'Файл пуст')
                return result

            # Parse headers
            headers = rows[0]
            column_map = {}
            for idx, header in enumerate(headers):
                if header:
                    normalized = self._normalize_column(str(header))
                    if normalized:
                        column_map[normalized] = idx

            if 'name' not in column_map or 'brand' not in column_map:
                result.add_error(1, 'Отсутствуют обязательные колонки: name, brand')
                return result

            # Process data rows
            for row_idx, row in enumerate(rows[1:], start=2):
                result.total_rows += 1
                self._process_row(row, column_map, row_idx, result)

            wb.close()

        except Exception as e:
            logger.exception(f'Error importing XLSX: {e}')
            result.add_error(0, f'Ошибка чтения файла: {e}')

        return result

    def import_csv(self, file: BinaryIO, encoding: str = 'utf-8') -> ImportResult:
        """Import products from CSV file."""
        result = ImportResult()

        try:
            # Read and decode
            content = file.read()
            try:
                text = content.decode(encoding)
            except UnicodeDecodeError:
                text = content.decode('cp1251')  # Fallback for Windows

            reader = csv.reader(io.StringIO(text))
            rows = list(reader)

            if not rows:
                result.add_error(0, 'Файл пуст')
                return result

            # Parse headers
            headers = rows[0]
            column_map = {}
            for idx, header in enumerate(headers):
                if header:
                    normalized = self._normalize_column(header)
                    if normalized:
                        column_map[normalized] = idx

            if 'name' not in column_map or 'brand' not in column_map:
                result.add_error(1, 'Отсутствуют обязательные колонки: name, brand')
                return result

            # Process data rows
            for row_idx, row in enumerate(rows[1:], start=2):
                result.total_rows += 1
                self._process_row(row, column_map, row_idx, result)

        except Exception as e:
            logger.exception(f'Error importing CSV: {e}')
            result.add_error(0, f'Ошибка чтения файла: {e}')

        return result

    @transaction.atomic
    def _process_row(self, row: tuple, column_map: dict, row_idx: int, result: ImportResult):
        """Process a single data row."""

        def get_value(field: str) -> str:
            idx = column_map.get(field)
            if idx is not None and idx < len(row):
                val = row[idx]
                return str(val).strip() if val else ''
            return ''

        # Get required fields
        name = get_value('name')
        brand = get_value('brand')

        if not name or not brand:
            result.add_error(row_idx, 'Пустое название или бренд')
            return

        # Parse optional fields
        is_own = self._parse_bool(get_value('is_own'))
        if is_own is None:
            is_own = True  # Default to own product

        product_data = {
            'brand': brand,
            'is_own': is_own,
            'product_type': get_value('product_type'),
            'packaging_type': self._parse_packaging(get_value('packaging_type')),
            'weight_grams': self._parse_int(get_value('weight_grams')),
            'caliber': get_value('caliber'),
            'has_pit': self._parse_bool(get_value('has_pit')),
            'variety': get_value('variety'),
            'notes': get_value('notes'),
        }

        # Create or update product
        try:
            product, created = Product.objects.update_or_create(
                name=name,
                brand=brand,
                defaults=product_data,
            )
            if created:
                result.products_created += 1
            else:
                result.products_updated += 1

        except Exception as e:
            result.add_error(row_idx, f'Ошибка создания товара: {e}')
            return

        # Process listings (URLs)
        url_fields = [
            ('url_ozon', 'ozon'),
            ('url_vkusvill', 'vkusvill'),
            ('url_perekrestok', 'perekrestok'),
            ('url_lavka', 'lavka'),
        ]

        for field, retailer_slug in url_fields:
            url = get_value(field)
            if url and url.startswith('http'):
                retailer = self._get_retailer(retailer_slug)
                if not retailer:
                    result.add_error(row_idx, f'Ретейлер {retailer_slug} не найден')
                    continue

                try:
                    listing, created = Listing.objects.update_or_create(
                        product=product,
                        retailer=retailer,
                        defaults={
                            'external_url': url,
                            'is_active': True,
                        }
                    )
                    if created:
                        result.listings_created += 1
                    else:
                        result.listings_updated += 1

                except Exception as e:
                    result.add_error(row_idx, f'Ошибка создания листинга {retailer_slug}: {e}')


class ProductExporter:
    """Export products to Excel/CSV."""

    HEADERS = [
        'Название',
        'Бренд',
        'Наш товар',
        'Тип продукта',
        'Упаковка',
        'Вес (г)',
        'Калибр',
        'Косточка',
        'Сорт',
        'Заметки',
        'URL Ozon',
        'URL ВкусВилл',
        'URL Перекрёсток',
        'URL Лавка',
    ]

    def export_xlsx(self, queryset=None) -> bytes:
        """Export products to Excel bytes."""
        if queryset is None:
            queryset = Product.objects.all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Товары'

        # Headers
        for col, header in enumerate(self.HEADERS, start=1):
            ws.cell(row=1, column=col, value=header)

        # Data
        for row_idx, product in enumerate(queryset.prefetch_related('listings__retailer'), start=2):
            listings = {l.retailer.slug: l.external_url for l in product.listings.all()}

            ws.cell(row=row_idx, column=1, value=product.name)
            ws.cell(row=row_idx, column=2, value=product.brand)
            ws.cell(row=row_idx, column=3, value='Да' if product.is_own else 'Нет')
            ws.cell(row=row_idx, column=4, value=product.product_type)
            ws.cell(row=row_idx, column=5, value=product.get_packaging_type_display() if product.packaging_type else '')
            ws.cell(row=row_idx, column=6, value=product.weight_grams)
            ws.cell(row=row_idx, column=7, value=product.caliber)
            ws.cell(row=row_idx, column=8, value='Да' if product.has_pit else ('Нет' if product.has_pit is False else ''))
            ws.cell(row=row_idx, column=9, value=product.variety)
            ws.cell(row=row_idx, column=10, value=product.notes)
            ws.cell(row=row_idx, column=11, value=listings.get('ozon', ''))
            ws.cell(row=row_idx, column=12, value=listings.get('vkusvill', ''))
            ws.cell(row=row_idx, column=13, value=listings.get('perekrestok', ''))
            ws.cell(row=row_idx, column=14, value=listings.get('lavka', ''))

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def export_csv(self, queryset=None) -> str:
        """Export products to CSV string."""
        if queryset is None:
            queryset = Product.objects.all()

        output = io.StringIO()
        writer = csv.writer(output)

        # Headers
        writer.writerow(self.HEADERS)

        # Data
        for product in queryset.prefetch_related('listings__retailer'):
            listings = {l.retailer.slug: l.external_url for l in product.listings.all()}

            writer.writerow([
                product.name,
                product.brand,
                'Да' if product.is_own else 'Нет',
                product.product_type,
                product.get_packaging_type_display() if product.packaging_type else '',
                product.weight_grams or '',
                product.caliber,
                'Да' if product.has_pit else ('Нет' if product.has_pit is False else ''),
                product.variety,
                product.notes,
                listings.get('ozon', ''),
                listings.get('vkusvill', ''),
                listings.get('perekrestok', ''),
                listings.get('lavka', ''),
            ])

        return output.getvalue()

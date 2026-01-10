"""
Excel export functionality for scraping data.
Provides clean, formatted exports for competitive intelligence analysis.
"""
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from .models import ManualImport, MonitoringGroup


# Styles
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
OWN_FILL = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
COMPETITOR_FILL = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
PRICE_UP_FONT = Font(color='DC2626', bold=True)
PRICE_DOWN_FONT = Font(color='16A34A', bold=True)
NEGATIVE_FILL = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
POSITIVE_FILL = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='E5E7EB'),
    right=Side(style='thin', color='E5E7EB'),
    top=Side(style='thin', color='E5E7EB'),
    bottom=Side(style='thin', color='E5E7EB'),
)


def format_price(value: Optional[Decimal]) -> str:
    """Format price value for display."""
    if value is None:
        return ''
    return f'{value:,.2f}'.replace(',', ' ')


def format_percentage(value: Optional[Decimal]) -> str:
    """Format percentage value."""
    if value is None:
        return ''
    sign = '+' if value > 0 else ''
    return f'{sign}{value:.1f}%'


class MonitoringExporter:
    """Export monitoring data to Excel with multiple sheets."""

    def __init__(self, user, period: Optional[datetime] = None):
        self.user = user
        self.period = period or datetime.now().replace(day=1).date()
        self.wb = Workbook()
        # Remove default sheet
        self.wb.remove(self.wb.active)

    def export(self) -> BytesIO:
        """Generate complete export with all sheets."""
        # Get all imports for this period
        imports = ManualImport.objects.filter(
            user=self.user,
            status=ManualImport.StatusChoices.COMPLETED,
        ).select_related('retailer', 'group').order_by('product_type', 'retailer__name')

        if self.period:
            imports = imports.filter(monitoring_period=self.period)

        # Create sheets
        self._create_summary_sheet(imports)
        self._create_price_monitoring_sheet(imports)
        self._create_own_products_sheet(imports.filter(product_type='own'))
        self._create_competitor_sheet(imports.filter(product_type='competitor'))
        self._create_review_insights_sheet(imports)

        # Save to buffer
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _create_summary_sheet(self, imports):
        """Summary sheet with key metrics."""
        ws = self.wb.create_sheet('Сводка')

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f'Мониторинг цен и отзывов - {self.period.strftime("%B %Y")}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Stats
        own_count = imports.filter(product_type='own').count()
        comp_count = imports.filter(product_type='competitor').count()

        own_with_price_up = imports.filter(
            product_type='own',
            price_change__gt=0
        ).count()
        own_with_price_down = imports.filter(
            product_type='own',
            price_change__lt=0
        ).count()

        comp_with_price_up = imports.filter(
            product_type='competitor',
            price_change__gt=0
        ).count()

        total_negative = sum(i.reviews_negative_count or 0 for i in imports.filter(product_type='own'))

        stats = [
            ('', ''),
            ('Статистика', ''),
            ('Наших товаров', own_count),
            ('Товаров конкурентов', comp_count),
            ('', ''),
            ('Изменения цен (наши)', ''),
            ('Повышение цены', own_with_price_up),
            ('Снижение цены', own_with_price_down),
            ('', ''),
            ('Изменения цен (конкуренты)', ''),
            ('Повышение цены', comp_with_price_up),
            ('', ''),
            ('Отзывы (наши товары)', ''),
            ('Негативных отзывов', total_negative),
        ]

        for row_idx, (label, value) in enumerate(stats, 3):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)
            if 'Статистика' in str(label) or 'Изменения' in str(label) or 'Отзывы' in str(label):
                ws.cell(row=row_idx, column=1).font = Font(bold=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

    def _create_price_monitoring_sheet(self, imports):
        """Price monitoring sheet with all products."""
        ws = self.wb.create_sheet('Мониторинг цен')

        headers = [
            'Тип', 'Название', 'Магазин', 'Текущая цена',
            'Предыдущая цена', 'Изменение', 'Изменение %',
            'Рейтинг', 'Отзывов', 'В наличии'
        ]

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER

        # Write data
        row = 2
        for imp in imports:
            product_type = 'Наш' if imp.product_type == 'own' else 'Конкурент'
            name = imp.custom_name or imp.product_title or imp.url[:50]
            retailer = imp.retailer.name if imp.retailer else '-'

            values = [
                product_type,
                name,
                retailer,
                float(imp.price_final) if imp.price_final else None,
                float(imp.price_previous) if imp.price_previous else None,
                float(imp.price_change) if imp.price_change else None,
                float(imp.price_change_pct) if imp.price_change_pct else None,
                float(imp.rating) if imp.rating else None,
                imp.reviews_count,
                'Да' if imp.in_stock else 'Нет' if imp.in_stock is False else '-'
            ]

            row_fill = OWN_FILL if imp.product_type == 'own' else COMPETITOR_FILL

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.fill = row_fill

                # Format price columns
                if col_idx in (4, 5) and value:
                    cell.number_format = '#,##0.00 ₽'
                elif col_idx == 6 and value:  # Price change
                    cell.number_format = '+#,##0.00 ₽;-#,##0.00 ₽'
                    if value > 0:
                        cell.font = PRICE_UP_FONT
                    elif value < 0:
                        cell.font = PRICE_DOWN_FONT
                elif col_idx == 7 and value:  # Percentage
                    cell.number_format = '+0.0%;-0.0%'
                    cell.value = value / 100  # Convert to decimal for percentage format
                    if value > 0:
                        cell.font = PRICE_UP_FONT
                    elif value < 0:
                        cell.font = PRICE_DOWN_FONT

            row += 1

        # Auto-fit columns
        column_widths = [10, 45, 15, 15, 15, 12, 12, 10, 10, 10]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

    def _create_own_products_sheet(self, imports):
        """Sheet for own products with detailed analysis."""
        ws = self.wb.create_sheet('Наши товары')

        headers = [
            'Название', 'Магазин', 'Цена', 'Изменение цены',
            'Рейтинг', 'Всего отзывов', 'Негативных', 'Нейтральных', 'Позитивных',
            'Проблемы (вкус)', 'Проблемы (упаковка)', 'Проблемы (качество)'
        ]

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER

        row = 2
        for imp in imports:
            name = imp.custom_name or imp.product_title or imp.url[:50]
            retailer = imp.retailer.name if imp.retailer else '-'

            # Extract topic insights
            insights = imp.review_insights or {}
            topics = insights.get('topics', {})

            taste_neg = topics.get('taste', {}).get('negative', 0)
            packaging_neg = topics.get('packaging', {}).get('negative', 0)
            quality_neg = topics.get('quality', {}).get('negative', 0)

            values = [
                name,
                retailer,
                float(imp.price_final) if imp.price_final else None,
                float(imp.price_change) if imp.price_change else None,
                float(imp.rating) if imp.rating else None,
                imp.reviews_count or 0,
                imp.reviews_negative_count or 0,
                imp.reviews_neutral_count or 0,
                imp.reviews_positive_count or 0,
                taste_neg,
                packaging_neg,
                quality_neg,
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = THIN_BORDER

                # Format price
                if col_idx == 3 and value:
                    cell.number_format = '#,##0.00 ₽'
                elif col_idx == 4 and value:
                    cell.number_format = '+#,##0.00 ₽;-#,##0.00 ₽'
                    if value > 0:
                        cell.font = PRICE_UP_FONT
                    elif value < 0:
                        cell.font = PRICE_DOWN_FONT
                elif col_idx == 7 and value > 0:  # Negative reviews
                    cell.fill = NEGATIVE_FILL
                elif col_idx in (10, 11, 12) and value > 0:  # Topic problems
                    cell.fill = NEGATIVE_FILL

            row += 1

        # Column widths
        column_widths = [40, 15, 12, 12, 10, 12, 12, 12, 12, 15, 15, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = 'A2'

    def _create_competitor_sheet(self, imports):
        """Sheet for competitor products with insights."""
        ws = self.wb.create_sheet('Конкуренты')

        headers = [
            'Название', 'Магазин', 'Цена', 'Изменение цены',
            'Рейтинг', 'Всего отзывов', 'Позитивных',
            'Плюсы (вкус)', 'Плюсы (упаковка)', 'Плюсы (качество)',
            'Примеры позитивных отзывов'
        ]

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER

        row = 2
        for imp in imports:
            name = imp.custom_name or imp.product_title or imp.url[:50]
            retailer = imp.retailer.name if imp.retailer else '-'

            # Extract topic insights
            insights = imp.review_insights or {}
            topics = insights.get('topics', {})

            taste_pos = topics.get('taste', {}).get('positive', 0)
            packaging_pos = topics.get('packaging', {}).get('positive', 0)
            quality_pos = topics.get('quality', {}).get('positive', 0)

            # Get sample positive reviews
            samples = []
            for topic in ['taste', 'packaging', 'quality']:
                topic_samples = topics.get(topic, {}).get('samples', [])
                samples.extend(topic_samples[:1])
            samples_text = ' | '.join(samples[:2]) if samples else ''

            values = [
                name,
                retailer,
                float(imp.price_final) if imp.price_final else None,
                float(imp.price_change) if imp.price_change else None,
                float(imp.rating) if imp.rating else None,
                imp.reviews_count or 0,
                imp.reviews_positive_count or 0,
                taste_pos,
                packaging_pos,
                quality_pos,
                samples_text[:200],
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = THIN_BORDER

                if col_idx == 3 and value:
                    cell.number_format = '#,##0.00 ₽'
                elif col_idx == 4 and value:
                    cell.number_format = '+#,##0.00 ₽;-#,##0.00 ₽'
                    if value > 0:
                        cell.font = PRICE_UP_FONT
                    elif value < 0:
                        cell.font = PRICE_DOWN_FONT
                elif col_idx in (8, 9, 10) and value > 0:  # Positive topics
                    cell.fill = POSITIVE_FILL
                elif col_idx == 11:  # Samples
                    cell.alignment = Alignment(wrap_text=True)

            row += 1

        # Column widths
        column_widths = [40, 15, 12, 12, 10, 12, 12, 12, 15, 15, 50]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = 'A2'

    def _create_review_insights_sheet(self, imports):
        """Sheet with detailed review insights for product development."""
        ws = self.wb.create_sheet('Инсайты из отзывов')

        # Section: Negative feedback on own products
        ws.merge_cells('A1:F1')
        ws['A1'] = 'Негативные отзывы на наши товары (для улучшения)'
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].fill = NEGATIVE_FILL

        headers = ['Товар', 'Тема', 'Кол-во', 'Пример отзыва']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

        row = 3
        own_imports = imports.filter(product_type='own')
        for imp in own_imports:
            insights = imp.review_insights or {}
            topics = insights.get('topics', {})
            name = imp.custom_name or imp.product_title or imp.url[:40]

            for topic_key, topic_name in [('taste', 'Вкус'), ('packaging', 'Упаковка'),
                                          ('quality', 'Качество'), ('price', 'Цена')]:
                topic_data = topics.get(topic_key, {})
                neg_count = topic_data.get('negative', 0)
                if neg_count > 0:
                    samples = topic_data.get('samples', [])
                    sample = samples[0][:200] if samples else ''

                    ws.cell(row=row, column=1, value=name).border = THIN_BORDER
                    ws.cell(row=row, column=2, value=topic_name).border = THIN_BORDER
                    ws.cell(row=row, column=3, value=neg_count).border = THIN_BORDER
                    sample_cell = ws.cell(row=row, column=4, value=sample)
                    sample_cell.border = THIN_BORDER
                    sample_cell.alignment = Alignment(wrap_text=True)
                    row += 1

        # Add spacing
        row += 2

        # Section: Positive feedback on competitors
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'Позитивные отзывы конкурентов (что перенять)'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = POSITIVE_FILL
        row += 1

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        row += 1

        comp_imports = imports.filter(product_type='competitor')
        for imp in comp_imports:
            insights = imp.review_insights or {}
            topics = insights.get('topics', {})
            name = imp.custom_name or imp.product_title or imp.url[:40]

            for topic_key, topic_name in [('taste', 'Вкус'), ('packaging', 'Упаковка'),
                                          ('quality', 'Качество')]:
                topic_data = topics.get(topic_key, {})
                pos_count = topic_data.get('positive', 0)
                if pos_count > 0:
                    samples = topic_data.get('samples', [])
                    sample = samples[0][:200] if samples else ''

                    ws.cell(row=row, column=1, value=name).border = THIN_BORDER
                    ws.cell(row=row, column=2, value=topic_name).border = THIN_BORDER
                    ws.cell(row=row, column=3, value=pos_count).border = THIN_BORDER
                    sample_cell = ws.cell(row=row, column=4, value=sample)
                    sample_cell.border = THIN_BORDER
                    sample_cell.alignment = Alignment(wrap_text=True)
                    row += 1

        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 60


def export_imports_to_excel(user, period=None) -> BytesIO:
    """
    Main export function.
    Returns a BytesIO buffer containing the Excel file.
    """
    exporter = MonitoringExporter(user, period)
    return exporter.export()


def export_single_import_to_excel(import_obj: ManualImport) -> BytesIO:
    """Export a single import with its reviews to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Данные товара'

    # Product info
    info = [
        ('Название', import_obj.custom_name or import_obj.product_title),
        ('URL', import_obj.url),
        ('Магазин', import_obj.retailer.name if import_obj.retailer else '-'),
        ('Тип', 'Наш товар' if import_obj.product_type == 'own' else 'Конкурент'),
        ('', ''),
        ('Цена', f'{import_obj.price_final} ₽' if import_obj.price_final else '-'),
        ('Предыдущая цена', f'{import_obj.price_previous} ₽' if import_obj.price_previous else '-'),
        ('Изменение цены', f'{import_obj.price_change:+.2f} ₽' if import_obj.price_change else '-'),
        ('Изменение %', f'{import_obj.price_change_pct:+.1f}%' if import_obj.price_change_pct else '-'),
        ('', ''),
        ('Рейтинг', str(import_obj.rating) if import_obj.rating else '-'),
        ('Всего отзывов', import_obj.reviews_count or 0),
        ('Позитивных', import_obj.reviews_positive_count),
        ('Нейтральных', import_obj.reviews_neutral_count),
        ('Негативных', import_obj.reviews_negative_count),
    ]

    for row_idx, (label, value) in enumerate(info, 1):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60

    # Reviews sheet
    if import_obj.reviews_data:
        ws_reviews = wb.create_sheet('Отзывы')

        headers = ['Рейтинг', 'Дата', 'Автор', 'Текст', 'Плюсы', 'Минусы']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_reviews.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

        for row_idx, review in enumerate(import_obj.reviews_data, 2):
            values = [
                review.get('rating', ''),
                review.get('date', ''),
                review.get('author', ''),
                review.get('text', ''),
                review.get('pros', ''),
                review.get('cons', ''),
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws_reviews.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical='top')

                # Highlight based on rating
                rating = review.get('rating', 3)
                if rating <= 2:
                    cell.fill = NEGATIVE_FILL
                elif rating >= 4:
                    cell.fill = POSITIVE_FILL

        ws_reviews.column_dimensions['A'].width = 10
        ws_reviews.column_dimensions['B'].width = 12
        ws_reviews.column_dimensions['C'].width = 15
        ws_reviews.column_dimensions['D'].width = 50
        ws_reviews.column_dimensions['E'].width = 30
        ws_reviews.column_dimensions['F'].width = 30
        ws_reviews.freeze_panes = 'A2'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
